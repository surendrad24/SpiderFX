# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------
import csv
import base64
import hashlib
import hmac
import html
import ipaddress
import json
import logging
import multiprocessing as mp
import os
import random
import re
import string
import struct
import time
from collections import Counter
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter
from urllib.parse import quote

import cherrypy
from cherrypy import _cperror

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

import secure

from sflib import SpiderFoot

from sfscan import startSpiderFootScanner

from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup

mp.set_start_method("spawn", force=True)


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''
    web_users = dict()

    def _newscan_sidebar_stats(self: 'SpiderFootWebUi') -> dict:
        """Build sidebar stats for HX-style new scan page."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        now = time.localtime()
        scans_this_month = 0
        for row in scans:
            try:
                created_ts = int(row[3])
            except Exception:
                continue
            c = time.localtime(created_ts)
            if c.tm_year == now.tm_year and c.tm_mon == now.tm_mon:
                scans_this_month += 1

        api_total = 0
        api_configured = 0
        for mod_name, mod_meta in self.config.get('__modules__', {}).items():
            if mod_name.startswith("sfp__stor_"):
                continue
            mod_cfg = self.config.get(mod_name, {})
            for opt_name in mod_meta.get('opts', {}).keys():
                if "api_key" not in opt_name.lower():
                    continue
                api_total += 1
                try:
                    if isinstance(mod_cfg, dict) and str(mod_cfg.get(opt_name, "")).strip():
                        api_configured += 1
                except Exception:
                    continue

        return {
            "scans_this_month": scans_this_month,
            "scan_limit": 999,
            "scan_duration_limit": "48 hours",
            "max_targets_per_scan": 1024,
            "data_retention_days": 180,
            "api_total": api_total,
            "api_configured": api_configured
        }

    def _passwd_file_path(self: 'SpiderFootWebUi') -> str:
        return SpiderFootHelpers.dataPath() + "/passwd"

    def _totp_file_path(self: 'SpiderFootWebUi') -> str:
        return SpiderFootHelpers.dataPath() + "/users_2fa.json"

    def _profiles_file_path(self: 'SpiderFootWebUi') -> str:
        return SpiderFootHelpers.dataPath() + "/spiderfx_profiles.json"

    def _load_passwd_map(self: 'SpiderFootWebUi') -> dict:
        users = {}
        passwd_file = self._passwd_file_path()
        if not os.path.isfile(passwd_file):
            return users
        try:
            with open(passwd_file, "r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line or ":" not in line:
                        continue
                    u = line.split(":", 1)[0].strip()
                    p = line.split(":", 1)[1].strip()
                    if u:
                        users[u] = p
        except Exception as e:
            self.log.error(f"Unable to load passwd map: {e}")
        return users

    def _save_passwd_map(self: 'SpiderFootWebUi', users: dict) -> bool:
        passwd_file = self._passwd_file_path()
        tmp_file = passwd_file + ".tmp"
        try:
            with open(tmp_file, "w", encoding="utf-8") as fp:
                for u in sorted(users.keys()):
                    fp.write(f"{u}:{users[u]}\n")
            os.replace(tmp_file, passwd_file)
            return True
        except Exception as e:
            self.log.error(f"Unable to save passwd map: {e}")
            return False

    def _is_hashed_secret(self: 'SpiderFootWebUi', secret: str) -> bool:
        if not isinstance(secret, str):
            return False
        return secret.startswith("pbkdf2_sha256$")

    def _hash_password(self: 'SpiderFootWebUi', password: str, iterations: int = 240000) -> str:
        salt = base64.b64encode(os.urandom(16)).decode("utf-8").rstrip("=")
        pwd_hash = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
        h64 = base64.b64encode(pwd_hash).decode("utf-8").rstrip("=")
        return f"pbkdf2_sha256${iterations}${salt}${h64}"

    def _verify_secret(self: 'SpiderFootWebUi', username: str, password: str, stored_secret: str) -> bool:
        if self._is_hashed_secret(stored_secret):
            try:
                _, iter_s, salt, hash_b64 = stored_secret.split("$", 3)
                iterations = int(iter_s)
                calc = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
                calc_b64 = base64.b64encode(calc).decode("utf-8").rstrip("=")
                return hmac.compare_digest(calc_b64, hash_b64)
            except Exception:
                return False

        if not hmac.compare_digest(str(stored_secret), str(password)):
            return False

        # Transparent one-time migration from plaintext to PBKDF2 hash.
        users = self._load_passwd_map()
        if username in users:
            users[username] = self._hash_password(password)
            if self._save_passwd_map(users):
                self.web_users = users
        return True

    def _load_totp_data(self: 'SpiderFootWebUi') -> dict:
        path = self._totp_file_path()
        if not os.path.isfile(path):
            return {"users": {}}
        try:
            with open(path, "r", encoding="utf-8") as fp:
                data = json.load(fp)
                if not isinstance(data, dict):
                    return {"users": {}}
                if not isinstance(data.get("users"), dict):
                    data["users"] = {}
                return data
        except Exception:
            return {"users": {}}

    def _save_totp_data(self: 'SpiderFootWebUi', data: dict) -> bool:
        path = self._totp_file_path()
        tmp = path + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as fp:
                json.dump(data, fp, indent=2)
            os.replace(tmp, path)
            return True
        except Exception as e:
            self.log.error(f"Unable to save 2FA data: {e}")
            return False

    def _load_profiles_data(self: 'SpiderFootWebUi') -> dict:
        path = self._profiles_file_path()
        if not os.path.isfile(path):
            return {"users": {}}
        try:
            with open(path, "r", encoding="utf-8") as fp:
                data = json.load(fp)
                if not isinstance(data, dict):
                    return {"users": {}}
                if not isinstance(data.get("users"), dict):
                    data["users"] = {}
                return data
        except Exception:
            return {"users": {}}

    def _save_profiles_data(self: 'SpiderFootWebUi', data: dict) -> bool:
        path = self._profiles_file_path()
        tmp = path + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as fp:
                json.dump(data, fp, indent=2)
            os.replace(tmp, path)
            return True
        except Exception as e:
            self.log.error(f"Unable to save profiles data: {e}")
            return False

    def _default_profile(self: 'SpiderFootWebUi', username: str) -> dict:
        return {
            "username": username,
            "full_name": "",
            "email": "",
            "company": "",
            "role": "",
            "use_case": "",
            "avatar_url": "",
            "api_keys": []
        }

    def _get_profile(self: 'SpiderFootWebUi', username: str) -> dict:
        profiles = self._load_profiles_data().get("users", {})
        profile = profiles.get(username, {})
        if not isinstance(profile, dict):
            profile = {}
        merged = self._default_profile(username)
        for key in ["full_name", "email", "company", "role", "use_case"]:
            merged[key] = str(profile.get(key, "") or "")
        merged["avatar_url"] = str(profile.get("avatar_url", "") or "")
        api_keys = profile.get("api_keys", [])
        if not isinstance(api_keys, list):
            api_keys = []
        cleaned_keys = []
        for item in api_keys:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name", "")).strip()
            value = str(item.get("value", "")).strip()
            if not name:
                continue
            cleaned_keys.append({"name": name, "value": value})
        merged["api_keys"] = cleaned_keys[:100]
        return merged

    def _generate_totp_secret(self: 'SpiderFootWebUi') -> str:
        return base64.b32encode(os.urandom(20)).decode("utf-8").rstrip("=")

    def _totp_at(self: 'SpiderFootWebUi', secret: str, for_time: int, period: int = 30, digits: int = 6) -> str:
        counter = int(for_time // period)
        padded = secret + "=" * ((8 - len(secret) % 8) % 8)
        key = base64.b32decode(padded, casefold=True)
        msg = struct.pack(">Q", counter)
        digest = hmac.new(key, msg, hashlib.sha1).digest()
        offset = digest[-1] & 0x0F
        code = struct.unpack(">I", digest[offset:offset + 4])[0] & 0x7FFFFFFF
        return str(code % (10 ** digits)).zfill(digits)

    def _verify_totp(self: 'SpiderFootWebUi', secret: str, otp: str, skew_steps: int = 1) -> bool:
        now = int(time.time())
        for step in range(-skew_steps, skew_steps + 1):
            candidate = self._totp_at(secret, now + (step * 30))
            if hmac.compare_digest(candidate, otp):
                return True
        return False

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')
        self.web_users = web_config.get('auth_users', {})

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:")
            .style_src("'self'", "'unsafe-inline'")
            .base_uri("'self'")
            .connect_src("'self'", "data:")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500
        tb = _cperror.format_exc()
        self.log.error(f"Unhandled web exception: {tb}")

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=tb)
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)

        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve info for scan."]).encode('utf-8')

        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve correlations for scan."]).encode('utf-8')

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.xlxs"
            else:
                fname = "SpiderFoot-correlations.xlxs"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, headings, sheetNameIndex=0)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)

            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.csv"
            else:
                fname = "SpiderFoot-correlations.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.xlsx"
            else:
                fname = scan_name + "-SpiderFoot.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.csv"
            else:
                fname = scan_name + "-SpiderFoot.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.json"
        else:
            fname = scan_name + "-SpiderFoot.json"

        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return SpiderFootHelpers.buildGraphJson([root], data)

        if not scan_name:
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        try:
            gexf_data = SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)
            if isinstance(gexf_data, str):
                return gexf_data.encode("utf-8")
            return gexf_data
        except Exception as e:
            self.log.error(f"scanviz gexf export failed: {e}", exc_info=True)
            cherrypy.response.status = 500
            return b"Unable to generate GEXF export."

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "-SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        try:
            gexf_data = SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)
            if isinstance(gexf_data, str):
                return gexf_data.encode("utf-8")
            return gexf_data
        except Exception as e:
            self.log.error(f"scanvizmulti gexf export failed: {e}", exc_info=True)
            cherrypy.response.status = 500
            return b"Unable to generate multi-scan GEXF export."

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhx(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return HX-like analytics for a scan as JSON."""
        return self._build_hx_analytics(id)

    def _build_hx_analytics(self: 'SpiderFootWebUi', id: str) -> dict:
        """Build HX-style analytics payload for a scan."""
        if not id:
            return {"error": "Scan ID not specified."}

        dbh = SpiderFootDb(self.config)
        scan = dbh.scanInstanceGet(id)
        if not scan:
            return {"error": "Invalid scan ID."}

        rows = dbh.scanResultEvent(id, filterFp=True)
        errors = dbh.scanErrors(id)
        corr_rows = dbh.scanCorrelationSummary(id, by="risk")

        modules = Counter()
        event_types = Counter()
        entities = Counter()
        timeline = Counter()
        ioc_types = Counter()
        ioc_examples = []
        ioc_seen = set()
        unique_values = set()
        unique_domains = set()
        unique_ips = set()
        attack_tactics = Counter()

        domain_re = re.compile(r'^(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}$', re.I)

        def attack_tactic_for_event(event_type: str) -> str:
            et = event_type.upper()
            if any(k in et for k in ["PASSWORD", "HASH", "COMPROMISED", "LEAK", "BREACH"]):
                return "Credential Access"
            if any(k in et for k in ["VULNERABILITY", "OPEN", "EXPOSED", "MISCONFIG"]):
                return "Initial Access"
            if any(k in et for k in ["MALICIOUS", "BLACKLISTED", "C2", "BOTNET", "TOR_SITE"]):
                return "Command and Control"
            if any(k in et for k in ["WEBSERVER", "SOFTWARE", "BANNER", "URL_", "TCP_PORT", "OS_", "FINGERPRINT"]):
                return "Discovery"
            if any(k in et for k in ["DOMAIN", "IP_ADDRESS", "INTERNET_NAME", "NETBLOCK", "EMAILADDR", "PHONE", "HUMAN_NAME", "USERNAME", "SOCIAL_MEDIA"]):
                return "Reconnaissance"
            return "Collection"

        for row in rows:
            generated = int(row[0])
            value = str(row[1]) if row[1] is not None else ""
            module = str(row[3]) if row[3] is not None else "unknown"
            event_type = str(row[11]) if row[11] is not None else str(row[4])

            modules[module] += 1
            event_types[event_type] += 1
            entities[value] += 1
            attack_tactics[attack_tactic_for_event(event_type)] += 1
            unique_values.add(value)
            timeline[time.strftime("%Y-%m-%d", time.localtime(generated))] += 1

            host = re.sub(r'^https?://', '', value.lower()).split('/')[0].split(':')[0]
            if domain_re.match(host):
                unique_domains.add(host)

            try:
                ipaddress.ip_address(host or value.strip())
                unique_ips.add(host or value.strip())
            except Exception:
                pass

            is_ioc = False
            if event_type.startswith("MALICIOUS_") or event_type.startswith("BLACKLISTED_"):
                is_ioc = True
            if "COMPROMISED" in event_type:
                is_ioc = True
            if event_type.startswith("VULNERABILITY_"):
                is_ioc = True
            if event_type in ["LEAKSITE_URL", "LEAKSITE_CONTENT", "BREACH_SITE", "TOR_SITE"]:
                is_ioc = True

            if is_ioc:
                ioc_types[event_type] += 1
                ex_key = f"{event_type}:{value}"
                if ex_key not in ioc_seen and len(ioc_examples) < 20:
                    ioc_seen.add(ex_key)
                    ioc_examples.append({"type": event_type, "value": value})

        corr_counts = {risk: 0 for risk in ["HIGH", "MEDIUM", "LOW", "INFO"]}
        for risk, total in corr_rows:
            corr_counts[str(risk)] = int(total)

        total_corr = sum(corr_counts.values())
        weighted = (
            corr_counts["HIGH"] * 40 +
            corr_counts["MEDIUM"] * 15 +
            corr_counts["LOW"] * 5 +
            corr_counts["INFO"] * 1
        )
        if total_corr > 0:
            risk_score = int(round((weighted / float(total_corr * 40)) * 100))
        else:
            ioc_total = sum(ioc_types.values())
            risk_score = min(100, int(ioc_total / 5))

        if risk_score >= 75:
            risk_level = "CRITICAL"
        elif risk_score >= 55:
            risk_level = "HIGH"
        elif risk_score >= 30:
            risk_level = "MEDIUM"
        elif risk_score > 0:
            risk_level = "LOW"
        else:
            risk_level = "INFO"

        def topn(counter_obj, key_name):
            return [{key_name: k, "count": v} for k, v in counter_obj.most_common(15)]

        return {
            "scan": {
                "id": id,
                "name": scan[0],
                "target": scan[1],
                "started": scan[3],
                "ended": scan[4],
                "status": scan[5]
            },
            "kpi": {
                "total_events": len(rows),
                "unique_entities": len(unique_values),
                "modules_used": len(modules),
                "event_types": len(event_types),
                "errors": len(errors),
                "domains": len(unique_domains),
                "ips": len(unique_ips),
                "ioc_events": sum(ioc_types.values())
            },
            "risk": {
                "score": risk_score,
                "level": risk_level,
                "correlations": corr_counts
            },
            "timeline": [{"date": d, "count": timeline[d]} for d in sorted(timeline.keys())],
            "modules": topn(modules, "module"),
            "event_types": topn(event_types, "event_type"),
            "attack_tactics": topn(attack_tactics, "tactic"),
            "entities": topn(entities, "value"),
            "ioc_types": topn(ioc_types, "event_type"),
            "ioc_examples": ioc_examples
        }

    @cherrypy.expose
    def scanhxreportpdf(self: 'SpiderFootWebUi', id: str) -> bytes:
        """Generate an executive HX-style PDF report for a scan."""
        data = self._build_hx_analytics(id)
        if data.get("error"):
            return self.error(data.get("error"))

        safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', str(data["scan"]["name"] or "SpiderFoot"))
        fname = f"{safe_name}-HX-Report.pdf"

        output = BytesIO()
        pdf = canvas.Canvas(output, pagesize=letter)
        width, height = letter
        y = height - 40

        def line(text: str, size: int = 10, bold: bool = False, indent: int = 40):
            nonlocal y
            if y < 50:
                pdf.showPage()
                y = height - 40
            font_name = "Helvetica-Bold" if bold else "Helvetica"
            pdf.setFont(font_name, size)
            pdf.drawString(indent, y, str(text))
            y -= (size + 4)

        line("SpiderFoot HX-Like Executive Report", 16, True)
        line(f"Scan Name: {data['scan']['name']}", 11, False)
        line(f"Target: {data['scan']['target']}", 11, False)
        line(f"Status: {data['scan']['status']}", 11, False)
        line(f"Started: {data['scan']['started']}", 11, False)
        line(f"Completed: {data['scan']['ended']}", 11, False)
        y -= 6

        line("Risk Summary", 13, True)
        line(f"Risk Score: {data['risk']['score']} / 100")
        line(f"Risk Level: {data['risk']['level']}")
        corr = data["risk"]["correlations"]
        line(f"Correlations: HIGH={corr.get('HIGH', 0)}  MEDIUM={corr.get('MEDIUM', 0)}  LOW={corr.get('LOW', 0)}  INFO={corr.get('INFO', 0)}")
        y -= 4

        line("Key Metrics", 13, True)
        kpi = data["kpi"]
        line(f"Total Events: {kpi['total_events']}")
        line(f"Unique Entities: {kpi['unique_entities']}")
        line(f"IOC Events: {kpi['ioc_events']}")
        line(f"Domains: {kpi['domains']}  IPs: {kpi['ips']}")
        line(f"Modules Used: {kpi['modules_used']}  Event Types: {kpi['event_types']}")
        line(f"Errors: {kpi['errors']}")
        y -= 4

        line("Top ATT&CK-Like Tactics", 13, True)
        for row in data.get("attack_tactics", [])[:8]:
            line(f"- {row['tactic']}: {row['count']}")

        y -= 4
        line("Top IOC Event Types", 13, True)
        for row in data.get("ioc_types", [])[:10]:
            line(f"- {row['event_type']}: {row['count']}")

        y -= 4
        line("Top Modules", 13, True)
        for row in data.get("modules", [])[:10]:
            line(f"- {row['module']}: {row['count']}")

        y -= 4
        line("Top Entities", 13, True)
        for row in data.get("entities", [])[:12]:
            value = str(row["value"])
            if len(value) > 90:
                value = value[:87] + "..."
            line(f"- {value}: {row['count']}")

        pdf.showPage()
        pdf.save()
        payload = output.getvalue()
        output.close()

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/pdf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return payload

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhxelement(self: 'SpiderFootWebUi', id: str, value: str) -> dict:
        """Return detailed HX-like insights for one element value in a scan."""
        if not id:
            return {"error": "Scan ID not specified."}
        if value is None:
            return {"error": "Element value not specified."}

        dbh = SpiderFootDb(self.config)
        scan = dbh.scanInstanceGet(id)
        if not scan:
            return {"error": "Invalid scan ID."}

        rows = dbh.scanResultEvent(id, filterFp=True)
        target_value = str(value)
        modules = Counter()
        event_types = Counter()
        parents = set()
        children = set()
        occurrences = 0
        first_seen = None
        last_seen = None
        risky = False

        for row in rows:
            generated = int(row[0])
            data_value = str(row[1]) if row[1] is not None else ""
            source_value = str(row[2]) if row[2] is not None else ""
            module = str(row[3]) if row[3] is not None else "unknown"
            event_type = str(row[11]) if row[11] is not None else str(row[4])

            if data_value == target_value:
                occurrences += 1
                modules[module] += 1
                event_types[event_type] += 1
                if source_value and source_value != data_value:
                    parents.add(source_value)
                if first_seen is None or generated < first_seen:
                    first_seen = generated
                if last_seen is None or generated > last_seen:
                    last_seen = generated

                et = event_type.upper()
                if et.startswith("MALICIOUS_") or et.startswith("VULNERABILITY_") or "COMPROMISED" in et or et.startswith("BLACKLISTED_"):
                    risky = True

            if source_value == target_value and data_value and data_value != source_value:
                children.add(data_value)

        def topn(counter_obj, key_name):
            return [{key_name: k, "count": v} for k, v in counter_obj.most_common(10)]

        return {
            "value": target_value,
            "occurrences": occurrences,
            "parent_count": len(parents),
            "child_count": len(children),
            "parents": list(parents)[:15],
            "children": list(children)[:15],
            "modules": topn(modules, "module"),
            "event_types": topn(event_types, "event_type"),
            "risky": risky,
            "first_seen": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(first_seen)) if first_seen else "",
            "last_seen": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(last_seen)) if last_seen else ""
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")

            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] Scan [{scanId}] failed: {e}")
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        sidebar = self._newscan_sidebar_stats()
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", sidebar=sidebar, version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')
        sidebar = self._newscan_sidebar_stats()

        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), sidebar=sidebar, version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show public spiderFX homepage.

        Returns:
            str: Homepage HTML
        """
        return self.fx()

    @cherrypy.expose
    def console(self: 'SpiderFootWebUi') -> str:
        """Show authenticated scan list console.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def fx(self: 'SpiderFootWebUi') -> str:
        """Public spiderFX homepage."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        scan_count = len(scans)
        active_count = 0
        for row in scans:
            if str(row[6]) in ["RUNNING", "STARTING", "STARTED"]:
                active_count += 1

        featured_tools = [
            "Amass", "Subfinder", "Nuclei", "Katana", "Naabu", "HTTPX",
            "FFUF", "CMSeeK", "TruffleHog", "PhoneInfoga", "Sherlock", "theHarvester"
        ]
        stats = self._newscan_sidebar_stats()
        templ = Template(filename='spiderfoot/templates/fx_home.tmpl', lookup=self.lookup)
        return templ.render(
            docroot=self.docroot,
            version=__version__,
            scan_count=scan_count,
            active_count=active_count,
            featured_tools=featured_tools,
            api_configured=stats.get("api_configured", 0),
            api_total=stats.get("api_total", 0)
        )

    @cherrypy.expose
    def signin(self: 'SpiderFootWebUi', next_url: str = None, error: str = None, success: str = None,
               otp_required: str = None) -> str:
        """Public spiderFX sign-in page."""
        if cherrypy.session.get("authenticated"):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/console")
        show_otp = bool(otp_required) or bool(cherrypy.session.get("pending_2fa_user"))
        pending_user = cherrypy.session.get("pending_2fa_user")
        templ = Template(filename='spiderfoot/templates/signin.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__, next_url=next_url, error=error,
                            success=success, show_otp=show_otp, pending_user=pending_user)

    @cherrypy.expose
    def signinsubmit(self: 'SpiderFootWebUi', username: str = "", password: str = "", otp: str = "",
                     next_url: str = None) -> None:
        """Sign-in handoff endpoint.

        spiderFX uses session auth backed by SpiderFoot passwd credentials.
        """
        username = (username or "").strip()
        password = (password or "").strip()
        otp = re.sub(r"\s+", "", otp or "")
        if len(username) < 2:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Please+enter+your+username")
        if len(password) < 1 and not cherrypy.session.get("pending_2fa_user"):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Please+enter+your+password")

        users = self._load_passwd_map()
        self.web_users = users
        pending_user = cherrypy.session.get("pending_2fa_user")
        pending_next = cherrypy.session.get("pending_2fa_next") or "/console"

        if pending_user:
            if username != pending_user:
                cherrypy.session.pop("pending_2fa_user", None)
                cherrypy.session.pop("pending_2fa_next", None)
                raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Please+retry+signin")

            totp = self._load_totp_data().get("users", {}).get(username, {})
            secret = str(totp.get("secret", ""))
            enabled = bool(totp.get("enabled", False))
            if not enabled or not secret:
                cherrypy.session.pop("pending_2fa_user", None)
                cherrypy.session.pop("pending_2fa_next", None)
                raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=2FA+is+not+configured")
            if not otp or not re.match(r"^\d{6}$", otp) or not self._verify_totp(secret, otp):
                nxt_q = quote(pending_next, safe="")
                raise cherrypy.HTTPRedirect(
                    f"{self.docroot}/signin?error=Invalid+2FA+code&otp_required=1&next_url={nxt_q}"
                )
            cherrypy.session.pop("pending_2fa_user", None)
            cherrypy.session.pop("pending_2fa_next", None)
            username = pending_user
        else:
            expected = users.get(username, None)
            if expected is None or not self._verify_secret(username, password, expected):
                nxt = "/console"
                if next_url and isinstance(next_url, str) and next_url.startswith("/"):
                    nxt = next_url
                nxt_q = quote(nxt, safe="")
                raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Invalid+credentials&next_url={nxt_q}")

            user_totp = self._load_totp_data().get("users", {}).get(username, {})
            if bool(user_totp.get("enabled", False)):
                nxt = "/console"
                if next_url and isinstance(next_url, str) and next_url.startswith("/"):
                    nxt = next_url
                cherrypy.session["pending_2fa_user"] = username
                cherrypy.session["pending_2fa_next"] = nxt
                nxt_q = quote(nxt, safe="")
                raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?otp_required=1&next_url={nxt_q}")

        cherrypy.session.regenerate()
        cherrypy.session["authenticated"] = True
        cherrypy.session["username"] = username

        target = f"{self.docroot}/console"
        chosen_next = pending_next if pending_user else next_url
        if chosen_next and isinstance(chosen_next, str) and chosen_next.startswith("/") and not chosen_next.startswith("//"):
            target = f"{self.docroot}{chosen_next}"
        elif next_url and isinstance(next_url, str) and next_url.startswith("/") and not next_url.startswith("//"):
            target = f"{self.docroot}{next_url}"
        raise cherrypy.HTTPRedirect(target)

    @cherrypy.expose
    def signout(self: 'SpiderFootWebUi') -> None:
        """End spiderFX session."""
        cherrypy.session.pop("authenticated", None)
        cherrypy.session.pop("username", None)
        cherrypy.session.pop("pending_2fa_user", None)
        cherrypy.session.pop("pending_2fa_next", None)
        cherrypy.session.regenerate()
        raise cherrypy.HTTPRedirect(f"{self.docroot}/signin")

    @cherrypy.expose
    def profile(self: 'SpiderFootWebUi', msg: str = None, error: str = None) -> str:
        """User profile and account settings page."""
        username = self._require_signed_in_user()
        profile = self._get_profile(username)
        totp_data = self._load_totp_data()
        user_totp = totp_data.get("users", {}).get(username, {})
        twofa_enabled = bool(user_totp.get("enabled", False))
        pending_secret = str(cherrypy.session.get("profile_2fa_secret", "") or "")
        templ = Template(filename='spiderfoot/templates/profile.tmpl', lookup=self.lookup)
        return templ.render(
            docroot=self.docroot,
            version=__version__,
            pageid="PROFILE",
            username=username,
            profile=profile,
            twofa_enabled=twofa_enabled,
            pending_2fa_secret=pending_secret,
            pending_2fa_otpauth=(
                f"otpauth://totp/spiderFX:{username}?secret={pending_secret}&issuer=spiderFX"
                if pending_secret else ""
            ),
            msg=msg,
            error=error
        )

    @cherrypy.expose
    def profileupdate(self: 'SpiderFootWebUi', full_name: str = "", email: str = "", company: str = "",
                      role: str = "", use_case: str = "", avatar_url: str = "") -> None:
        """Update signed-in user profile fields."""
        username = self._require_signed_in_user()
        full_name = (full_name or "").strip()
        email = (email or "").strip().lower()
        company = (company or "").strip()
        role = (role or "").strip()
        use_case = (use_case or "").strip()
        avatar_url = (avatar_url or "").strip()

        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Please+enter+a+valid+email")
        if avatar_url and not re.match(r"^https?://[^\s]+$", avatar_url):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Avatar+URL+must+start+with+http(s)://")

        profiles_data = self._load_profiles_data()
        users = profiles_data.setdefault("users", {})
        old_profile = users.get(username, {})
        old_api_keys = []
        if isinstance(old_profile, dict) and isinstance(old_profile.get("api_keys"), list):
            old_api_keys = old_profile.get("api_keys", [])
        users[username] = {
            "full_name": full_name,
            "email": email,
            "company": company,
            "role": role,
            "use_case": use_case,
            "avatar_url": avatar_url,
            "api_keys": old_api_keys
        }
        if not self._save_profiles_data(profiles_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Unable+to+save+profile")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=Profile+updated")

    @cherrypy.expose
    def profileapikeys(self: 'SpiderFootWebUi', api_keys_json: str = "") -> None:
        """Update signed-in user personal API key vault."""
        username = self._require_signed_in_user()
        raw = (api_keys_json or "").strip()
        if not raw:
            parsed = []
        else:
            try:
                parsed = json.loads(raw)
            except Exception:
                raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Invalid+API+vault+payload")
            if not isinstance(parsed, list):
                raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Invalid+API+vault+payload")

        cleaned = []
        for item in parsed[:100]:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name", "")).strip()
            value = str(item.get("value", "")).strip()
            if not name:
                continue
            if len(name) > 100 or len(value) > 5000:
                continue
            cleaned.append({"name": name, "value": value})

        profiles_data = self._load_profiles_data()
        users = profiles_data.setdefault("users", {})
        old = users.get(username, {})
        if not isinstance(old, dict):
            old = {}
        users[username] = {
            "full_name": str(old.get("full_name", "") or ""),
            "email": str(old.get("email", "") or ""),
            "company": str(old.get("company", "") or ""),
            "role": str(old.get("role", "") or ""),
            "use_case": str(old.get("use_case", "") or ""),
            "avatar_url": str(old.get("avatar_url", "") or ""),
            "api_keys": cleaned
        }
        if not self._save_profiles_data(profiles_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Unable+to+save+API+vault")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=API+vault+updated")

    @cherrypy.expose
    def changepassword(self: 'SpiderFootWebUi', current_password: str = "", new_password: str = "",
                       confirm_password: str = "") -> None:
        """Change password for signed-in user."""
        username = self._require_signed_in_user()
        current_password = (current_password or "").strip()
        new_password = (new_password or "").strip()
        confirm_password = (confirm_password or "").strip()

        if len(current_password) < 1:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Please+enter+current+password")
        if len(new_password) < 10:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=New+password+must+be+at+least+10+characters")
        if new_password != confirm_password:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=New+passwords+do+not+match")
        if new_password == current_password:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=New+password+must+be+different")

        users = self._load_passwd_map()
        stored = users.get(username)
        if not stored or not self._verify_secret(username, current_password, stored):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Current+password+is+incorrect")

        users[username] = self._hash_password(new_password)
        if not self._save_passwd_map(users):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Unable+to+update+password")

        self.web_users = users
        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=Password+updated")

    @cherrypy.expose
    def profile2fasetup(self: 'SpiderFootWebUi') -> None:
        """Start self-service 2FA setup for signed-in user."""
        username = self._require_signed_in_user()
        totp_data = self._load_totp_data()
        user_totp = totp_data.get("users", {}).get(username, {})
        if bool(user_totp.get("enabled", False)):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=2FA+already+enabled")
        cherrypy.session["profile_2fa_secret"] = self._generate_totp_secret()
        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=Scan+the+QR+and+verify+code+to+enable+2FA")

    @cherrypy.expose
    def profile2faenable(self: 'SpiderFootWebUi', otp: str = "") -> None:
        """Finalize self-service 2FA setup by verifying OTP."""
        username = self._require_signed_in_user()
        otp = re.sub(r"\s+", "", otp or "")
        secret = str(cherrypy.session.get("profile_2fa_secret", "") or "")
        if not secret:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=No+pending+2FA+setup")
        if not re.match(r"^\d{6}$", otp) or not self._verify_totp(secret, otp):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Invalid+2FA+code")

        totp_data = self._load_totp_data()
        totp_data.setdefault("users", {})[username] = {"enabled": True, "secret": secret}
        if not self._save_totp_data(totp_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Unable+to+enable+2FA")
        cherrypy.session.pop("profile_2fa_secret", None)
        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=2FA+enabled")

    @cherrypy.expose
    def profile2facancel(self: 'SpiderFootWebUi') -> None:
        """Cancel pending self-service 2FA setup."""
        self._require_signed_in_user()
        cherrypy.session.pop("profile_2fa_secret", None)
        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=2FA+setup+canceled")

    @cherrypy.expose
    def profile2fadisable(self: 'SpiderFootWebUi', current_password: str = "") -> None:
        """Disable self-service 2FA for signed-in user after password check."""
        username = self._require_signed_in_user()
        current_password = (current_password or "").strip()
        if len(current_password) < 1:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Please+enter+your+password")

        users = self._load_passwd_map()
        stored = users.get(username)
        if not stored or not self._verify_secret(username, current_password, stored):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Current+password+is+incorrect")

        totp_data = self._load_totp_data()
        if username in totp_data.get("users", {}):
            totp_data["users"][username]["enabled"] = False
        if not self._save_totp_data(totp_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?error=Unable+to+disable+2FA")
        cherrypy.session.pop("profile_2fa_secret", None)
        raise cherrypy.HTTPRedirect(f"{self.docroot}/profile?msg=2FA+disabled")

    @cherrypy.expose
    def signup(self: 'SpiderFootWebUi', success: str = None, error: str = None) -> str:
        """Public spiderFX sign-up page."""
        templ = Template(filename='spiderfoot/templates/signup.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__, success=success, error=error)

    @cherrypy.expose
    def signupsubmit(self: 'SpiderFootWebUi', full_name: str = "", email: str = "", company: str = "",
                     role: str = "", use_case: str = "", username: str = "", password: str = "",
                     confirm_password: str = "") -> None:
        """Store sign-up interest for spiderFX onboarding."""
        full_name = (full_name or "").strip()
        email = (email or "").strip().lower()
        company = (company or "").strip()
        role = (role or "").strip()
        use_case = (use_case or "").strip()
        username = (username or "").strip()
        password = (password or "").strip()
        confirm_password = (confirm_password or "").strip()

        if len(full_name) < 2:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Please+enter+your+name")

        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Please+enter+a+valid+email")
        if not re.match(r"^[a-zA-Z0-9_.-]{3,32}$", username):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Choose+a+valid+username+(3-32,+letters/numbers/._-)")
        if len(password) < 10:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Password+must+be+at+least+10+characters")
        if password != confirm_password:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Passwords+do+not+match")

        csv_file = SpiderFootHelpers.dataPath() + "/spiderfx_signups.csv"
        try:
            new_file = False
            try:
                with open(csv_file, "r", encoding="utf-8"):
                    pass
            except FileNotFoundError:
                new_file = True

            with open(csv_file, "a", encoding="utf-8", newline="") as fp:
                writer = csv.writer(fp)
                if new_file:
                    writer.writerow(["timestamp", "username", "full_name", "email", "company", "role", "use_case"])
                writer.writerow([int(time.time()), username, full_name, email, company, role, use_case])
        except Exception as e:
            self.log.error(f"Unable to save spiderFX signup: {e}")
            # Do not block account creation if analytics logging fails.

        users = self._load_passwd_map()
        if username in users:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Username+already+exists")
        users[username] = self._hash_password(password)
        if not self._save_passwd_map(users):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signup?error=Could+not+create+login+account")
        self.web_users = users

        profiles_data = self._load_profiles_data()
        users_profiles = profiles_data.setdefault("users", {})
        users_profiles[username] = {
            "full_name": full_name,
            "email": email,
            "company": company,
            "role": role,
            "use_case": use_case,
            "avatar_url": "",
            "api_keys": []
        }
        if not self._save_profiles_data(profiles_data):
            self.log.error(f"Unable to persist profile data for user {username}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?success=Account+created.+Please+sign+in")

    def _require_admin(self: 'SpiderFootWebUi') -> str:
        user = cherrypy.session.get("username")
        if user != "admin":
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Admin+access+required")
        return user

    def _require_signed_in_user(self: 'SpiderFootWebUi') -> str:
        user = cherrypy.session.get("username")
        if not user or not cherrypy.session.get("authenticated"):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/signin?error=Please+sign+in")
        return str(user)

    @cherrypy.expose
    def usermgmt(self: 'SpiderFootWebUi', msg: str = None, error: str = None, show_secret_for: str = None) -> str:
        """Admin user management panel."""
        self._require_admin()
        users = self._load_passwd_map()
        self.web_users = users
        totp_data = self._load_totp_data()
        totp_users = totp_data.get("users", {})
        rows = []
        for username in sorted(users.keys()):
            u2fa = totp_users.get(username, {})
            secret = str(u2fa.get("secret", ""))
            rows.append({
                "username": username,
                "hashed": self._is_hashed_secret(users[username]),
                "twofa_enabled": bool(u2fa.get("enabled", False)),
                "secret": secret if show_secret_for == username else "",
                "otpauth": f"otpauth://totp/spiderFX:{username}?secret={secret}&issuer=spiderFX"
                if show_secret_for == username and secret else ""
            })
        templ = Template(filename='spiderfoot/templates/usermgmt.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, pageid="USERMGMT", version=__version__, rows=rows,
                            msg=msg, error=error, show_secret_for=show_secret_for)

    @cherrypy.expose
    def usercreate(self: 'SpiderFootWebUi', username: str = "", password: str = "") -> None:
        """Create a new login user (admin only)."""
        self._require_admin()
        username = (username or "").strip()
        password = (password or "").strip()
        if not re.match(r"^[a-zA-Z0-9_.-]{3,32}$", username):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Invalid+username")
        if len(password) < 10:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Password+must+be+at+least+10+chars")
        users = self._load_passwd_map()
        if username in users:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Username+already+exists")
        users[username] = self._hash_password(password)
        if not self._save_passwd_map(users):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Unable+to+save+user")
        self.web_users = users
        raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?msg=User+created")

    @cherrypy.expose
    def userdelete(self: 'SpiderFootWebUi', username: str = "") -> None:
        """Delete a login user (admin only)."""
        self._require_admin()
        username = (username or "").strip()
        if username == "admin":
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Cannot+delete+admin")
        users = self._load_passwd_map()
        if username not in users:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=User+not+found")
        users.pop(username, None)
        if not self._save_passwd_map(users):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Unable+to+delete+user")
        self.web_users = users
        totp_data = self._load_totp_data()
        totp_data.setdefault("users", {}).pop(username, None)
        self._save_totp_data(totp_data)
        raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?msg=User+deleted")

    @cherrypy.expose
    def user2faenable(self: 'SpiderFootWebUi', username: str = "") -> None:
        """Enable or rotate TOTP 2FA for a user (admin only)."""
        self._require_admin()
        username = (username or "").strip()
        users = self._load_passwd_map()
        if username not in users:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=User+not+found")
        totp_data = self._load_totp_data()
        secret = self._generate_totp_secret()
        totp_data.setdefault("users", {})[username] = {"enabled": True, "secret": secret}
        if not self._save_totp_data(totp_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Unable+to+save+2FA")
        raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?msg=2FA+enabled&show_secret_for={quote(username, safe='')}")

    @cherrypy.expose
    def user2fadisable(self: 'SpiderFootWebUi', username: str = "") -> None:
        """Disable TOTP 2FA for a user (admin only)."""
        self._require_admin()
        username = (username or "").strip()
        totp_data = self._load_totp_data()
        if username in totp_data.get("users", {}):
            totp_data["users"][username]["enabled"] = False
        if not self._save_totp_data(totp_data):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?error=Unable+to+update+2FA")
        raise cherrypy.HTTPRedirect(f"{self.docroot}/usermgmt?msg=2FA+disabled")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def fxhealth(self: 'SpiderFootWebUi') -> dict:
        """Public health and capability summary for spiderFX landing pages."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        active = 0
        completed = 0
        for row in scans:
            status = str(row[6])
            if status in ["RUNNING", "STARTING", "STARTED"]:
                active += 1
            if status in ["FINISHED", "ABORTED", "ERROR-FAILED"]:
                completed += 1
        modules_total = len(self.config.get("__modules__", {}))
        integrations = []
        for mod in self.config.get("__modules__", {}).keys():
            if mod.startswith("sfp_tool_"):
                integrations.append(mod.replace("sfp_tool_", ""))

        return {
            "brand": "spiderFX",
            "version": __version__,
            "scans_total": len(scans),
            "scans_active": active,
            "scans_completed": completed,
            "modules_total": modules_total,
            "integrations_total": len(integrations),
            "integrations_sample": sorted(integrations)[:20]
        }

    @cherrypy.expose
    def investigate(self: 'SpiderFootWebUi') -> str:
        """Show investigations-oriented view of notable correlations."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        rows = []
        for scan in scans[:30]:
            scan_id = scan[0]
            scan_name = scan[1]
            status = scan[6]
            if str(status) not in ["FINISHED", "ABORTED", "ERROR-FAILED"]:
                continue
            corr = dbh.scanCorrelationList(scan_id)
            for item in corr:
                risk = str(item[3])
                if risk not in ["HIGH", "MEDIUM", "LOW"]:
                    continue
                rows.append({
                    "scan_id": scan_id,
                    "scan_name": scan_name,
                    "title": item[1],
                    "risk": risk,
                    "count": item[7]
                })
        risk_weight = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
        rows.sort(key=lambda r: (risk_weight.get(r["risk"], 0), int(r["count"])), reverse=True)

        templ = Template(filename='spiderfoot/templates/investigate.tmpl', lookup=self.lookup)
        return templ.render(pageid='INVESTIGATE', docroot=self.docroot, rows=rows[:200], version=__version__)

    @cherrypy.expose
    def monitor(self: 'SpiderFootWebUi') -> str:
        """Show monitor-style change summary across targets."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        by_target = {}
        for row in scans:
            target = row[2]
            by_target.setdefault(target, []).append(row)

        monitor_rows = []
        for target, target_scans in by_target.items():
            target_scans.sort(key=lambda x: int(x[3] or 0), reverse=True)
            if len(target_scans) < 2:
                continue
            latest = target_scans[0]
            prev = target_scans[1]
            latest_unique = dbh.scanResultEventUnique(latest[0], filterFp=True)
            prev_unique = dbh.scanResultEventUnique(prev[0], filterFp=True)

            latest_vals = set([str(x[0]) for x in latest_unique])
            prev_vals = set([str(x[0]) for x in prev_unique])
            added = latest_vals - prev_vals
            removed = prev_vals - latest_vals

            latest_events = dbh.scanResultEvent(latest[0], filterFp=True)
            prev_events = dbh.scanResultEvent(prev[0], filterFp=True)

            def risky_count(events):
                cnt = 0
                for e in events:
                    et = str(e[11])
                    if et.startswith("MALICIOUS_") or et.startswith("VULNERABILITY_") or "COMPROMISED" in et:
                        cnt += 1
                return cnt

            latest_risky = risky_count(latest_events)
            prev_risky = risky_count(prev_events)

            monitor_rows.append({
                "target": target,
                "latest_scan": latest[1],
                "latest_scan_id": latest[0],
                "previous_scan": prev[1],
                "added": len(added),
                "removed": len(removed),
                "risky_delta": latest_risky - prev_risky,
                "latest_elements": len(latest_unique),
                "previous_elements": len(prev_unique)
            })

        monitor_rows.sort(key=lambda r: (r["added"] + abs(r["risky_delta"])), reverse=True)

        templ = Template(filename='spiderfoot/templates/monitor.tmpl', lookup=self.lookup)
        return templ.render(pageid='MONITOR', docroot=self.docroot, rows=monitor_rows[:100], version=__version__)

    @cherrypy.expose
    def apidocs(self: 'SpiderFootWebUi') -> str:
        """Show API docs page."""
        templ = Template(filename='spiderfoot/templates/apidocs.tmpl', lookup=self.lookup)
        return templ.render(pageid='APIDOCS', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        templ = Template(filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        """List all modules.

        Returns:
            list: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret

        modinfo.sort()

        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all correlation rules.

        Returns:
            list: list of correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        rules = self.config['__correlationrules__']
        if not rules:
            return ret

        for r in rules:
            ret.append({
                'id': r['id'],
                'name': r['meta']['name'],
                'descr': r['meta']['description'],
                'risk': r['meta']['risk'],
            })

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: SpiderFoot version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Unrecognised target type."]).encode('utf-8')

            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        # Swap the globalscantable for the database handler
        dbh = SpiderFootDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            if usecase == 'DeepRecon':
                # Opinionated high-coverage recon preset focused on the custom
                # ProjectDiscovery/Kali-style modules integrated on this host.
                deep_recon_modules = [
                    "sfp_dnsresolve",
                    "sfp_dnsbrute",
                    "sfp_spider",
                    "sfp_torch",
                    "sfp_ahmia",
                    "sfp_onionsearchengine",
                    "sfp_tool_subfinder",
                    "sfp_tool_amass",
                    "sfp_tool_dnsx",
                    "sfp_tool_naabu",
                    "sfp_tool_httpx",
                    "sfp_tool_katana",
                    "sfp_tool_gau",
                    "sfp_tool_waybackurls",
                    "sfp_tool_tlsx",
                    "sfp_tool_nuclei",
                    "sfp_tool_uncover",
                    "sfp_tool_dnstwist",
                    "sfp_tool_cmseek",
                    "sfp_tool_trufflehog",
                    "sfp_tool_searchsploit",
                    "sfp_tool_ffuf",
                    "sfp_tool_maigret",
                    "sfp_tool_h8mail",
                    "sfp_tool_bbot",
                    "sfp_tool_theharvester",
                    "sfp_tool_holehe",
                    "sfp_tool_sherlock",
                    "sfp_tool_phoneinfoga",
                    "sfp_tool_metagoofil",
                    "sfp_osint_portals",
                    "sfp_mcp_server",
                ]
                for mod_name, opt_name in [
                    ("sfp_haveibeenpwned", "api_key"),
                    ("sfp_greynoise", "api_key"),
                    ("sfp_intelx", "api_key"),
                ]:
                    mod_cfg = self.config.get(mod_name, {})
                    if isinstance(mod_cfg, dict) and str(mod_cfg.get(opt_name, "")).strip():
                        deep_recon_modules.append(mod_name)
                modlist = [m for m in deep_recon_modules if m in self.config['__modules__']]
            else:
                for mod in self.config['__modules__']:
                    if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                        modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()

        # Delete the stdout module in case it crept in
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata
